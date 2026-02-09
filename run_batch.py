import os
import base64
from pathlib import Path
from datetime import datetime

from openai import OpenAI
from openpyxl import Workbook, load_workbook


# ============== 配置区 ==============
IMAGES_DIR = Path("./SchadenFotos")
EXCEL_PATH = Path("./Foto2Text_SchadenTest.xlsx")  # 用你现有的；没有也行，会自动创建
SHEET_NAME = "AI_Schaden_Test"  # 如果你的表叫别的，改这里
MODEL_NAME = "gpt-4.1-mini"     # 你也可以换成更强的模型（成本更高）
AI_MODEL_LABEL = f"OpenAI:{MODEL_NAME}"  # 写入 Excel 的“AI_Modell”列

PROMPT_DE = """Du bist ein Assistent zur Schadenbeschreibung.
Beschreibe den sichtbaren Schaden auf dem Foto objektiv, klar und detailliert.

Regeln:
- Nur sichtbare Fakten beschreiben.
- Keine Vermutungen über Ursache, Schuld oder Kosten.
- Wenn etwas nicht erkennbar ist: "nicht beurteilbar".

Bitte liefere die Ausgabe in Stichpunkten mit diesen Feldern:
- Objekt/Teil
- Schadenart
- Position am Objekt
- Schweregrad (leicht/mittel/stark/nicht beurteilbar)
- Funktionsbeeinträchtigung (falls erkennbar)
- Sonstige Hinweise
"""
# ====================================


def image_to_data_url(image_path: Path) -> str:
    ext = image_path.suffix.lower().lstrip(".")
    if ext == "jpg":
        ext = "jpeg"
    mime = f"image/{ext}"
    b64 = base64.b64encode(image_path.read_bytes()).decode("utf-8")
    return f"data:{mime};base64,{b64}"


def ensure_workbook(path: Path, sheet_name: str):
    """
    如果 Excel 不存在：创建一个带标准表头的文件。
    如果存在：直接打开。
    """
    headers = [
        "Bild_ID",
        "Objektkategorie",
        "Schwierigkeitsgrad",
        "Beschreibung_Bild",
        "AI_Modell",
        "AI_Beschreibung",
        "Richtigkeit_1-5",
        "Detailgrad_1-5",
        "Objektivitaet_1-5",
        "Umgang_mit_Unsicherheit_1-5",
        "Gesamtpunktzahl",
        "Kommentar",
        "Timestamp",
    ]

    if path.exists():
        wb = load_workbook(path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    wb.save(path)
    return wb, ws


def call_openai_vision(client: OpenAI, model: str, prompt: str, image_path: Path) -> str:
    data_url = image_to_data_url(image_path)

    resp = client.responses.create(
        model=model,
        input=[{
            "role": "user",
            "content": [
                {"type": "input_text", "text": prompt},
                {"type": "input_image", "image_url": data_url},
            ],
        }],
    )

    # SDK 提供 output_text 快捷属性（更稳）
    return (resp.output_text or "").strip()


def main():
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY 环境变量未设置")

    if not IMAGES_DIR.exists():
        raise RuntimeError(f"图片文件夹不存在：{IMAGES_DIR.resolve()}")

    client = OpenAI(api_key=api_key)
    wb, ws = ensure_workbook(EXCEL_PATH, SHEET_NAME)

    image_files = sorted([p for p in IMAGES_DIR.iterdir()
                          if p.suffix.lower() in [".jpg", ".jpeg", ".png", ".webp"]])

    if not image_files:
        raise RuntimeError(f"在 {IMAGES_DIR.resolve()} 没找到图片（jpg/png/webp）")

    for img_path in image_files:
        bild_id = img_path.stem  # 用文件名当 Bild_ID，例如 car_001

        print(f"▶ Processing: {img_path.name}")
        try:
            ai_text = call_openai_vision(client, MODEL_NAME, PROMPT_DE, img_path)
        except Exception as e:
            ai_text = f"[ERROR] {type(e).__name__}: {e}"

        # 你可以先把 Objektkategorie / Schwierigkeitsgrad 留空，后面手动补
        row = [
            bild_id,
            "",  # Objektkategorie
            "",  # Schwierigkeitsgrad
            "",  # Beschreibung_Bild (你可手动写一句“图片内容概述”)
            AI_MODEL_LABEL,
            ai_text,
            "", "", "", "", "", "",  # 评分 & 评论先留空
            datetime.now().isoformat(timespec="seconds"),
        ]
        ws.append(row)

    wb.save(EXCEL_PATH)
    print(f"\n✅ Done. Results saved to: {EXCEL_PATH.resolve()}")


if __name__ == "__main__":
    main()
